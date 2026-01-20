"""Generate a new quotation Excel file from a template and JSON data."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy

TEMPLATE_PATH = Path("templates/Formato de cotizaciones Antartida y Altavolt.xlsx")
DATA_PATH = Path("data/cotizacion.json")
TEMPLATE_SHEET_NAME = "Lomas Country Temixco"

HEADER_LABELS = {
    "cliente": "Cliente:",
    "direccion": "Dirección:",
    "telefono": "Teléfono:",
    "fecha": "Fecha del presupuesto",
}

TABLE_HEADER_LABEL = "DESCRIPCIÓN"


def load_payload() -> dict:
    with DATA_PATH.open("r", encoding="utf-8") as file:
        return json.load(file)


def find_cell_with_text(sheet: Worksheet, text: str) -> Optional[Cell]:
    """Locate a cell whose text matches exactly the provided label."""
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == text:
                return cell
    return None


def find_cell_startswith(sheet: Worksheet, prefix: str) -> Optional[Cell]:
    """Locate a cell whose text starts with the provided prefix."""
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().startswith(prefix):
                return cell
    return None


def find_table_header(sheet: Worksheet) -> Tuple[int, Dict[str, int]]:
    """Find the header row for the concepts table and return column mapping."""
    for row in sheet.iter_rows():
        header_map: Dict[str, int] = {}
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            label = cell.value.strip().upper()
            if label == TABLE_HEADER_LABEL:
                header_map["descripcion"] = cell.column
            if label == "UNIDADES":
                header_map["unidades"] = cell.column
            if label == "PRECIO":
                header_map["precio_unitario"] = cell.column
            if label == "TOTAL":
                header_map["total"] = cell.column
        if "descripcion" in header_map:
            return row[0].row, header_map
    raise ValueError("No se encontró la fila de encabezado de la tabla de conceptos.")


def clear_existing_concepts(sheet: Worksheet, start_row: int, columns: Iterable[int]) -> int:
    """Clear only the content of existing concepts without touching styles."""
    row = start_row
    cleared_rows = 0
    while True:
        values = [sheet.cell(row=row, column=col).value for col in columns]
        if all(value is None for value in values):
            break
        for col in columns:
            cell = sheet.cell(row=row, column=col)
            if cell.data_type != "f":
                cell.value = None
        row += 1
        cleared_rows += 1
    return cleared_rows


def copy_row_style(sheet: Worksheet, source_row: int, target_row: int) -> None:
    """Copy row height and cell styles from the source row."""
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for col in range(1, sheet.max_column + 1):
        source_cell = sheet.cell(row=source_row, column=col)
        target_cell = sheet.cell(row=target_row, column=col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)


def write_concepts(sheet: Worksheet, concepts: Iterable[dict], header_row: int, header_map: dict) -> None:
    """Write concepts under the table header, inserting rows if needed."""
    data_row = header_row + 1
    target_columns = [header_map[key] for key in header_map]
    cleared_rows = clear_existing_concepts(sheet, data_row, target_columns)

    base_row = data_row
    for index, concept in enumerate(concepts):
        row = data_row + index
        if index >= cleared_rows:
            sheet.insert_rows(row)
            copy_row_style(sheet, base_row, row)

        descripcion_col = header_map["descripcion"]
        descripcion_cell = sheet.cell(row=row, column=descripcion_col)
        descripcion_cell.value = concept.get("descripcion")
        alignment = copy(descripcion_cell.alignment)
        alignment.wrap_text = True
        descripcion_cell.alignment = alignment

        unidades_col = header_map.get("unidades")
        precio_col = header_map.get("precio_unitario")
        total_col = header_map.get("total")

        unidades = concept.get("unidades", 0)
        precio_unitario = concept.get("precio_unitario", 0)

        if unidades_col:
            sheet.cell(row=row, column=unidades_col).value = unidades
        if precio_col:
            sheet.cell(row=row, column=precio_col).value = precio_unitario
        if total_col:
            total_cell = sheet.cell(row=row, column=total_col)
            if total_cell.data_type != "f":
                total_cell.value = float(unidades or 0) * float(precio_unitario or 0)


def main() -> None:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"No se encontró la plantilla: {TEMPLATE_PATH}")
    if not DATA_PATH.exists():
        raise FileNotFoundError(f"No se encontró el catálogo: {DATA_PATH}")

    payload = load_payload()
    proyecto = payload.get("proyecto", {})
    conceptos = payload.get("conceptos", [])

    workbook = load_workbook(TEMPLATE_PATH)
    if TEMPLATE_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"No se encontró la hoja plantilla: {TEMPLATE_SHEET_NAME}")

    template_sheet = workbook[TEMPLATE_SHEET_NAME]
    new_sheet = workbook.copy_worksheet(template_sheet)
    new_sheet.title = proyecto.get("folio", "Cotizacion")
    workbook.remove(template_sheet)

    cliente_cell = find_cell_with_text(new_sheet, HEADER_LABELS["cliente"])
    if cliente_cell:
        cliente_cell.value = f"{HEADER_LABELS['cliente']} {proyecto.get('cliente', '')}"

    direccion_cell = find_cell_with_text(new_sheet, HEADER_LABELS["direccion"])
    if direccion_cell:
        direccion_cell.value = f"{HEADER_LABELS['direccion']} {proyecto.get('direccion', '')}"

    telefono_cell = find_cell_with_text(new_sheet, HEADER_LABELS["telefono"])
    if telefono_cell:
        telefono_cell.value = f"{HEADER_LABELS['telefono']} {proyecto.get('telefono', '')}"

    fecha_cell = find_cell_with_text(new_sheet, HEADER_LABELS["fecha"])
    if fecha_cell:
        fecha_cell.offset(column=1).value = proyecto.get("fecha", "")

    titulo_cell = find_cell_startswith(new_sheet, "Presupuesto")
    if titulo_cell:
        titulo_cell.value = f"Presupuesto {proyecto.get('folio', '')}"

    header_row, header_map = find_table_header(new_sheet)
    write_concepts(new_sheet, conceptos, header_row, header_map)

    output_path = Path(f"dist/Cotizacion_{proyecto.get('folio', 'Generada')}.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


if __name__ == "__main__":
    main()
